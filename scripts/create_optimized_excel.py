import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Ruta del archivo original
archivo_original = r'C:\Users\aleja\OneDrive\Desktop\Control de Gastos SUPERNOVA.xlsx'
archivo_optimizado = r'C:\Users\aleja\OneDrive\Desktop\Control de Gastos SUPERNOVA - OPTIMIZADO.xlsx'

print("Leyendo archivo original...")

# Leer el archivo original
xl = pd.ExcelFile(archivo_original)

# Leer hojas principales
gastos_detallados_raw = pd.read_excel(xl, sheet_name='GASTOS DETALLADOS', header=None)
nomina_raw = pd.read_excel(xl, sheet_name='NOMINA', header=None)
resumen_raw = pd.read_excel(xl, sheet_name='RESUMEN', header=None)

print("Procesando datos...")

# ===============================
# PROCESAR GASTOS DETALLADOS
# ===============================
# Encontrar la fila de encabezados (buscar "Código Obra")
header_row = None
for idx, row in gastos_detallados_raw.iterrows():
    if 'Código Obra' in str(row.values) or 'C.digo Obra' in str(row.values):
        header_row = idx
        break

if header_row is not None:
    # Extraer encabezados y datos
    headers = gastos_detallados_raw.iloc[header_row].fillna('Sin_nombre')
    gastos_data = gastos_detallados_raw.iloc[header_row+1:].reset_index(drop=True)
    gastos_data.columns = headers

    # Limpiar y normalizar
    gastos_clean = gastos_data[gastos_data.iloc[:, 1].notna()].copy()

    # Renombrar columnas clave
    col_map = {}
    for i, col in enumerate(gastos_clean.columns):
        col_str = str(col)
        if 'digo' in col_str or 'Código' in col_str:
            col_map[col] = 'Código_Obra'
        elif 'Fecha' in col_str and 'Solicitud' in col_str:
            col_map[col] = 'Fecha_Solicitud'
        elif 'O.C' in col_str or 'OC' in col_str:
            col_map[col] = 'Orden_Compra'
        elif 'Estatus' in col_str and 'Pago' in col_str:
            col_map[col] = 'Estatus_Pago'
        elif 'Obra' == col_str.strip():
            col_map[col] = 'Nombre_Obra'
        elif 'Monto' in col_str:
            col_map[col] = 'Monto_Neto'
        elif 'Beneficiario' in col_str:
            col_map[col] = 'Beneficiario'
        elif 'Solicita' in col_str:
            col_map[col] = 'Solicitante'
        elif 'Comentarios' in col_str and i < 10:
            col_map[col] = 'Comentarios'
        elif 'Seguimiento' in col_str:
            col_map[col] = 'Estatus_Entrega'

    gastos_clean = gastos_clean.rename(columns=col_map)

    # Seleccionar columnas principales
    cols_to_keep = ['Código_Obra', 'Fecha_Solicitud', 'Orden_Compra', 'Estatus_Pago',
                    'Nombre_Obra', 'Monto_Neto', 'Beneficiario', 'Solicitante',
                    'Comentarios', 'Estatus_Entrega']

    gastos_final = gastos_clean[[col for col in cols_to_keep if col in gastos_clean.columns]].copy()

    # Limpiar datos
    if 'Monto_Neto' in gastos_final.columns:
        gastos_final['Monto_Neto'] = pd.to_numeric(gastos_final['Monto_Neto'], errors='coerce')

    if 'Fecha_Solicitud' in gastos_final.columns:
        gastos_final['Fecha_Solicitud'] = pd.to_datetime(gastos_final['Fecha_Solicitud'], errors='coerce')

    # Agregar columna de categoría
    gastos_final.insert(5, 'Tipo_Comprobante', 'Fiscal')
    gastos_final.insert(6, 'Categoría', 'General')

    # Agregar ID único
    gastos_final.insert(0, 'ID_Gasto', range(1, len(gastos_final) + 1))
else:
    gastos_final = pd.DataFrame()

# ===============================
# PROCESAR NÓMINA
# ===============================
# Buscar todas las tablas de nómina
nomina_records = []

for idx, row in nomina_raw.iterrows():
    if 'NOMBRE' in str(row.values) and 'PUESTO' in str(row.values):
        # Encontramos un encabezado de nómina
        header_idx = idx
        # Buscar la fila con el período
        periodo_row = nomina_raw.iloc[header_idx - 1] if header_idx > 0 else None
        periodo = str(periodo_row.values[2]) if periodo_row is not None else 'Desconocido'

        # Leer datos hasta la siguiente tabla
        data_idx = header_idx + 1
        while data_idx < len(nomina_raw):
            current_row = nomina_raw.iloc[data_idx]

            # Verificar si llegamos al final de esta tabla
            if pd.isna(current_row.iloc[1]) or 'NOMBRE' in str(current_row.values):
                break

            # Extraer datos
            nombre = current_row.iloc[1]
            puesto = current_row.iloc[2]
            sueldo = pd.to_numeric(current_row.iloc[3], errors='coerce')
            viaticos = pd.to_numeric(current_row.iloc[4], errors='coerce')
            nomina_total = pd.to_numeric(current_row.iloc[5], errors='coerce')
            dias_laborados = pd.to_numeric(current_row.iloc[6], errors='coerce')

            # Buscar descuentos y total pago en columnas posteriores
            descuentos = 0
            total_pago = nomina_total

            for i in range(7, min(len(current_row), 18)):
                val = pd.to_numeric(current_row.iloc[i], errors='coerce')
                if pd.notna(val) and val > 0:
                    if i == 12:  # Columna de descuentos
                        descuentos = val
                    elif i == 16:  # Columna de pago total
                        total_pago = val

            if pd.notna(nombre) and nombre != 'NOMBRE':
                nomina_records.append({
                    'Período': periodo,
                    'Nombre_Empleado': nombre,
                    'Puesto': puesto,
                    'Sueldo_Base': sueldo,
                    'Viáticos': viaticos,
                    'Nómina_Total': nomina_total,
                    'Días_Laborados': dias_laborados,
                    'Descuentos': descuentos,
                    'Total_Pagar': total_pago
                })

            data_idx += 1

nomina_final = pd.DataFrame(nomina_records)

# Agregar ID único
if not nomina_final.empty:
    nomina_final.insert(0, 'ID_Registro', range(1, len(nomina_final) + 1))

# ===============================
# CREAR CATÁLOGOS
# ===============================
# Catálogo de Obras
if not gastos_final.empty and 'Código_Obra' in gastos_final.columns:
    obras_unique = gastos_final[['Código_Obra', 'Nombre_Obra']].drop_duplicates()
    obras_unique = obras_unique[obras_unique['Código_Obra'].notna()]

    catalogo_obras = pd.DataFrame({
        'Código_Obra': obras_unique['Código_Obra'].values,
        'Nombre_Obra': obras_unique['Nombre_Obra'].values,
        'Presupuesto_Total': [153876208.21 if x == 1 else 0 for x in obras_unique['Código_Obra'].values],
        'Estatus': 'Activa',
        'Fecha_Inicio': '2022-11-01',
        'Responsable': 'Alejandro Elizalde'
    })
else:
    catalogo_obras = pd.DataFrame({
        'Código_Obra': [1],
        'Nombre_Obra': ['ZF TOLUCA'],
        'Presupuesto_Total': [153876208.21],
        'Estatus': ['Activa'],
        'Fecha_Inicio': ['2022-11-01'],
        'Responsable': ['Alejandro Elizalde']
    })

# Catálogo de Empleados
if not nomina_final.empty:
    empleados_unique = nomina_final[['Nombre_Empleado', 'Puesto']].drop_duplicates()

    catalogo_empleados = pd.DataFrame({
        'ID_Empleado': range(1, len(empleados_unique) + 1),
        'Nombre_Completo': empleados_unique['Nombre_Empleado'].values,
        'Puesto': empleados_unique['Puesto'].values,
        'Estatus': 'Activo',
        'Fecha_Ingreso': ''
    })
else:
    catalogo_empleados = pd.DataFrame({
        'ID_Empleado': [],
        'Nombre_Completo': [],
        'Puesto': [],
        'Estatus': [],
        'Fecha_Ingreso': []
    })

# Catálogo de Proveedores/Beneficiarios
if not gastos_final.empty and 'Beneficiario' in gastos_final.columns:
    proveedores_unique = gastos_final['Beneficiario'].dropna().unique()

    catalogo_proveedores = pd.DataFrame({
        'ID_Proveedor': range(1, len(proveedores_unique) + 1),
        'Nombre_Proveedor': proveedores_unique,
        'RFC': '',
        'Contacto': '',
        'Teléfono': '',
        'Tipo': 'Proveedor'
    })
else:
    catalogo_proveedores = pd.DataFrame({
        'ID_Proveedor': [],
        'Nombre_Proveedor': [],
        'RFC': [],
        'Contacto': [],
        'Teléfono': [],
        'Tipo': []
    })

print("Creando archivo Excel optimizado...")

# ===============================
# CREAR NUEVO ARCHIVO EXCEL
# ===============================
wb = Workbook()

# Estilos
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_alignment = Alignment(horizontal='left', vertical='center')
number_format = '#,##0.00'
date_format = 'DD/MM/YYYY'

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ===============================
# HOJA 1: DASHBOARD
# ===============================
ws_dashboard = wb.active
ws_dashboard.title = "Dashboard"

# Título
ws_dashboard['A1'] = 'CONTROL DE GASTOS SUPERNOVA'
ws_dashboard['A1'].font = Font(name='Calibri', size=18, bold=True, color='0066CC')
ws_dashboard.merge_cells('A1:F1')

ws_dashboard['A2'] = f'Última actualización: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws_dashboard['A2'].font = Font(size=10, italic=True)
ws_dashboard.merge_cells('A2:F2')

# KPIs
ws_dashboard['A4'] = 'INDICADORES PRINCIPALES'
ws_dashboard['A4'].font = Font(size=14, bold=True)
ws_dashboard.merge_cells('A4:F4')

kpis = [
    ['Métrica', 'Valor', 'Estado'],
    ['Total Obras Activas', f'=COUNTIF(Catálogo_Obras!D:D,"Activa")', ''],
    ['Gastos Total a la Fecha', f'=SUM(Gastos!G:G)', ''],
    ['Presupuesto Total', f'=SUM(Catálogo_Obras!C:C)', ''],
    ['% Ejecutado', '=B7/B8', ''],
    ['Utilidad Estimada', '=B8-B7', ''],
    ['Empleados Activos', f'=COUNTIF(Catálogo_Empleados!D:D,"Activo")', ''],
    ['Nómina Mes Actual', '', '']
]

for i, row in enumerate(kpis, start=5):
    for j, value in enumerate(row, start=1):
        cell = ws_dashboard.cell(row=i, column=j, value=value)
        if i == 5:  # Header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        else:
            if j == 2:  # Valores numéricos
                cell.number_format = number_format
            cell.alignment = data_alignment
        cell.border = thin_border

# Ajustar anchos
ws_dashboard.column_dimensions['A'].width = 25
ws_dashboard.column_dimensions['B'].width = 20
ws_dashboard.column_dimensions['C'].width = 15

# ===============================
# HOJA 2: CATÁLOGO DE OBRAS
# ===============================
ws_obras = wb.create_sheet("Catálogo_Obras")

# Escribir encabezados
headers_obras = list(catalogo_obras.columns)
for col_num, header in enumerate(headers_obras, 1):
    cell = ws_obras.cell(row=1, column=col_num, value=header.replace('_', ' '))
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Escribir datos
for r_idx, row in enumerate(dataframe_to_rows(catalogo_obras, index=False, header=False), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws_obras.cell(row=r_idx, column=c_idx, value=value)
        cell.border = thin_border
        if c_idx == 3:  # Presupuesto
            cell.number_format = number_format

# Ajustar anchos
ws_obras.column_dimensions['A'].width = 12
ws_obras.column_dimensions['B'].width = 20
ws_obras.column_dimensions['C'].width = 18
ws_obras.column_dimensions['D'].width = 12
ws_obras.column_dimensions['E'].width = 15
ws_obras.column_dimensions['F'].width = 20

# ===============================
# HOJA 3: CATÁLOGO DE EMPLEADOS
# ===============================
ws_empleados = wb.create_sheet("Catálogo_Empleados")

if not catalogo_empleados.empty:
    headers_emp = list(catalogo_empleados.columns)
    for col_num, header in enumerate(headers_emp, 1):
        cell = ws_empleados.cell(row=1, column=col_num, value=header.replace('_', ' '))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for r_idx, row in enumerate(dataframe_to_rows(catalogo_empleados, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_empleados.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

    ws_empleados.column_dimensions['A'].width = 12
    ws_empleados.column_dimensions['B'].width = 30
    ws_empleados.column_dimensions['C'].width = 25
    ws_empleados.column_dimensions['D'].width = 12
    ws_empleados.column_dimensions['E'].width = 15

# ===============================
# HOJA 4: CATÁLOGO DE PROVEEDORES
# ===============================
ws_proveedores = wb.create_sheet("Catálogo_Proveedores")

if not catalogo_proveedores.empty:
    headers_prov = list(catalogo_proveedores.columns)
    for col_num, header in enumerate(headers_prov, 1):
        cell = ws_proveedores.cell(row=1, column=col_num, value=header.replace('_', ' '))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for r_idx, row in enumerate(dataframe_to_rows(catalogo_proveedores, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_proveedores.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

    ws_proveedores.column_dimensions['A'].width = 12
    ws_proveedores.column_dimensions['B'].width = 35
    ws_proveedores.column_dimensions['C'].width = 15
    ws_proveedores.column_dimensions['D'].width = 20
    ws_proveedores.column_dimensions['E'].width = 15
    ws_proveedores.column_dimensions['F'].width = 15

# ===============================
# HOJA 5: GASTOS
# ===============================
ws_gastos = wb.create_sheet("Gastos")

if not gastos_final.empty:
    headers_gastos = list(gastos_final.columns)
    for col_num, header in enumerate(headers_gastos, 1):
        cell = ws_gastos.cell(row=1, column=col_num, value=header.replace('_', ' '))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for r_idx, row in enumerate(dataframe_to_rows(gastos_final, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_gastos.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

            # Formato especial para montos
            if headers_gastos[c_idx-1] == 'Monto_Neto':
                cell.number_format = number_format

            # Formato para fechas
            if headers_gastos[c_idx-1] == 'Fecha_Solicitud' and value:
                cell.number_format = date_format

    # Ajustar anchos
    ws_gastos.column_dimensions['A'].width = 10  # ID
    ws_gastos.column_dimensions['B'].width = 12  # Código Obra
    ws_gastos.column_dimensions['C'].width = 15  # Fecha
    ws_gastos.column_dimensions['D'].width = 15  # OC
    ws_gastos.column_dimensions['E'].width = 12  # Estatus
    ws_gastos.column_dimensions['F'].width = 20  # Nombre Obra
    ws_gastos.column_dimensions['G'].width = 15  # Tipo
    ws_gastos.column_dimensions['H'].width = 15  # Categoría
    ws_gastos.column_dimensions['I'].width = 15  # Monto
    ws_gastos.column_dimensions['J'].width = 30  # Beneficiario
    ws_gastos.column_dimensions['K'].width = 20  # Solicitante
    ws_gastos.column_dimensions['L'].width = 35  # Comentarios
    ws_gastos.column_dimensions['M'].width = 15  # Estatus Entrega

    # Agregar validación de datos para Estatus_Pago
    dv_estatus = DataValidation(type="list", formula1='"Pagado,Pendiente,Cancelado"', allow_blank=True)
    dv_estatus.error = 'Valor no válido'
    dv_estatus.errorTitle = 'Error de entrada'
    ws_gastos.add_data_validation(dv_estatus)
    dv_estatus.add(f'E2:E{len(gastos_final) + 10}')

    # Validación para Tipo_Comprobante
    dv_tipo = DataValidation(type="list", formula1='"Fiscal,No Fiscal"', allow_blank=True)
    ws_gastos.add_data_validation(dv_tipo)
    dv_tipo.add(f'G2:G{len(gastos_final) + 10}')

# ===============================
# HOJA 6: NÓMINA
# ===============================
ws_nomina = wb.create_sheet("Nómina")

if not nomina_final.empty:
    headers_nomina = list(nomina_final.columns)
    for col_num, header in enumerate(headers_nomina, 1):
        cell = ws_nomina.cell(row=1, column=col_num, value=header.replace('_', ' '))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for r_idx, row in enumerate(dataframe_to_rows(nomina_final, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws_nomina.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

            # Formato para montos
            if c_idx in [4, 5, 6, 8, 9]:  # Columnas de dinero
                cell.number_format = number_format

    # Ajustar anchos
    ws_nomina.column_dimensions['A'].width = 10
    ws_nomina.column_dimensions['B'].width = 30
    ws_nomina.column_dimensions['C'].width = 30
    ws_nomina.column_dimensions['D'].width = 25
    ws_nomina.column_dimensions['E'].width = 15
    ws_nomina.column_dimensions['F'].width = 15
    ws_nomina.column_dimensions['G'].width = 18
    ws_nomina.column_dimensions['H'].width = 15
    ws_nomina.column_dimensions['I'].width = 15
    ws_nomina.column_dimensions['J'].width = 15

# ===============================
# HOJA 7: RESUMEN MENSUAL
# ===============================
ws_mensual = wb.create_sheet("Resumen_Mensual")

# Crear tabla de resumen mensual
ws_mensual['A1'] = 'RESUMEN DE GASTOS MENSUALES'
ws_mensual['A1'].font = Font(size=14, bold=True)
ws_mensual.merge_cells('A1:P1')

headers_mensual = ['Código Obra', 'Obra', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                   'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre', 'Total', 'Presupuesto']

for col_num, header in enumerate(headers_mensual, 1):
    cell = ws_mensual.cell(row=2, column=col_num, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Agregar una fila de ejemplo
ws_mensual['A3'] = 1
ws_mensual['B3'] = 'ZF TOLUCA'
for col in range(3, 15):
    cell = ws_mensual.cell(row=3, column=col, value=0)
    cell.number_format = number_format
    cell.border = thin_border

ws_mensual['O3'] = '=SUM(C3:N3)'
ws_mensual['O3'].number_format = number_format
ws_mensual['P3'] = 153876208.21
ws_mensual['P3'].number_format = number_format

# Ajustar anchos
ws_mensual.column_dimensions['A'].width = 12
ws_mensual.column_dimensions['B'].width = 20
for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
    ws_mensual.column_dimensions[col_letter].width = 14

print("Guardando archivo...")

# Guardar archivo
wb.save(archivo_optimizado)

print(f"\n{'='*60}")
print("ARCHIVO CREADO EXITOSAMENTE")
print(f"{'='*60}")
print(f"\nRuta: {archivo_optimizado}")
print(f"\nESTADÍSTICAS:")
print(f"  - Gastos procesados: {len(gastos_final) if not gastos_final.empty else 0}")
print(f"  - Registros de nómina: {len(nomina_final) if not nomina_final.empty else 0}")
print(f"  - Obras catalogadas: {len(catalogo_obras)}")
print(f"  - Empleados: {len(catalogo_empleados)}")
print(f"  - Proveedores: {len(catalogo_proveedores)}")
print(f"\nHOJAS CREADAS:")
print("  1. Dashboard - Indicadores principales")
print("  2. Catálogo_Obras - Lista maestra de obras")
print("  3. Catálogo_Empleados - Lista de personal")
print("  4. Catálogo_Proveedores - Lista de beneficiarios")
print("  5. Gastos - Registro completo con validaciones")
print("  6. Nómina - Registro normalizado de pagos")
print("  7. Resumen_Mensual - Vista por meses")
print(f"\n{'='*60}")
