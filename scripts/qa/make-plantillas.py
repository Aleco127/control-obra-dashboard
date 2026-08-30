# Genera docs/plantillas/opus-ejemplo.xlsx y neodata-ejemplo.xlsx (US-242) con la estructura de los exportados
# a Excel de OPUS y Neodata (partidas sin cantidad, conceptos con clave, unidad, cantidad, precio e importe),
# y los CSV equivalentes en scripts/qa/fixtures para la prueba de node.
import csv, io, os
from openpyxl import Workbook
from openpyxl.styles import Font
R = os.path.join(os.path.dirname(__file__), '..', '..')
os.makedirs(os.path.join(R, 'docs', 'plantillas'), exist_ok=True)
os.makedirs(os.path.join(R, 'scripts', 'qa', 'fixtures'), exist_ok=True)

# Datos: casa habitación (tomados de la cotización OPUS de Supernova para David Juárez, redondeados)
DATA = [
    ('A', 'PRELIMINARES', '', None, None),
    ('A.01', 'Limpieza y trazo del terreno con equipo topográfico', 'm2', 800.00, 18.50),
    ('A.02', 'Excavación a mano en cepas material tipo II hasta 1.50 m', 'm3', 96.00, 285.00),
    ('A.03', 'Relleno compactado con material de banco en capas de 20 cm', 'm3', 42.00, 310.00),
    ('B', 'CIMENTACIÓN', '', None, None),
    ('B.01', 'Plantilla de concreto f\'c=100 kg/cm2 de 5 cm de espesor', 'm2', 64.00, 145.00),
    ('B.02', 'Zapata corrida de concreto f\'c=250 kg/cm2 armada, incluye cimbra', 'm3', 28.50, 4850.00),
    ('B.03', 'Cadena de desplante 15x30 cm concreto f\'c=250 armex 15-30-4', 'ml', 118.00, 425.00),
    ('C', 'ESTRUCTURA', '', None, None),
    ('C.01', 'Muro de block de concreto 15x20x40 asentado con mortero 1:4', 'm2', 420.00, 385.00),
    ('C.02', 'Castillo K-1 15x15 cm concreto f\'c=250 armex 15-15-4', 'ml', 260.00, 295.00),
    ('C.03', 'Losa maciza 12 cm concreto f\'c=250 armada, incluye cimbra', 'm2', 210.00, 1650.00),
    ('D', 'ACABADOS', '', None, None),
    ('D.01', 'Aplanado de mortero cemento-arena 1:4 acabado fino', 'm2', 840.00, 165.00),
    ('D.02', 'Piso de porcelanato 60x60 asentado con adhesivo', 'm2', 190.00, 690.00),
    ('D.03', 'Pintura vinílica dos manos sobre aplanado', 'm2', 840.00, 78.00),
]

def escribir(nombre, encabezados, titulo, cuerpo):
    wb = Workbook(); ws = wb.active; ws.title = 'Presupuesto'
    ws.append([titulo]); ws['A1'].font = Font(bold=True, size=12)
    ws.append(['Obra: Casa habitación 20x40 · Chihuahua, Chih.'])
    ws.append([])
    ws.append(encabezados)
    for c in ws[4]: c.font = Font(bold=True)
    filas = []
    for cve, desc, uni, cant, pu in DATA:
        if cant is None:
            sub = sum(c * p for k, d, u, c, p in DATA if c is not None and k.startswith(cve + '.'))
            fila = cuerpo(cve, desc, '', '', '', round(sub, 2), True)
        else:
            fila = cuerpo(cve, desc, uni, cant, pu, round(cant * pu, 2), False)
        ws.append(fila); filas.append(fila)
    total = sum(c * p for k, d, u, c, p in DATA if c is not None)
    ws.append([]); ws.append(cuerpo('', 'TOTAL', '', '', '', round(total, 2), True))
    for col, w in zip('ABCDEFG', [10, 62, 8, 12, 14, 16, 8]): ws.column_dimensions[col].width = w
    wb.save(os.path.join(R, 'docs', 'plantillas', nombre + '.xlsx'))
    with io.open(os.path.join(R, 'scripts', 'qa', 'fixtures', nombre + '.csv'), 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f); w.writerow([titulo]); w.writerow([]); w.writerow(encabezados)
        for fila in filas: w.writerow(fila)
        w.writerow(cuerpo('', 'TOTAL', '', '', '', round(total, 2), True))
    return round(total, 2)

t1 = escribir('opus-ejemplo', ['Código', 'Concepto', 'Unidad', 'Cantidad', 'P. Unitario', 'Importe'],
              'OPUS · Presupuesto de obra', lambda k, d, u, c, p, i, part: [k, d, u, c, p, i])
t2 = escribir('neodata-ejemplo', ['Clave', 'Descripción', 'Unidad', 'Cantidad', 'Precio', 'Importe', '%'],
              'Neodata · Presupuesto', lambda k, d, u, c, p, i, part: [k, d, u, c, p, i, ''])
print('plantillas listas; total', t1, t2)

# --- Formato real de OPUS 24 (skill /opus-budget-direct) ------------------------------
# El exportador de OPUS pone el nombre del proyecto en B5, los encabezados en la fila 6 y usa
# claves NN-XXX para la partida y NN-XXX-NNN para el concepto. Un catalogo de concurso sale con
# los precios en cero. Se generan las dos variantes para las pruebas.
OPUS24 = [
    ('01-PRE', 'PRELIMINARES Y DEMOLICIONES', '', None, None),
    ('01-PRE-005', 'PROTECCION DE AREAS ADYACENTES PREVIA A DEMOLICIONES A BASE DE PLASTICO CAL. 700 Y LONA', 'LOTE', 1.00, 850.00),
    ('01-PRE-010', 'RETIRO DE RECUBRIMIENTO EXISTENTE EN MUROS DE ADOBE DE FACHADA FRONTAL POR MEDIOS MANUALES', 'M2', 40.00, 95.00),
    ('01-PRE-020', 'DEMOLICION DE FIRME DE CONCRETO EXISTENTE EN COCHERA DE E=10 CM POR MEDIOS MECANICOS', 'M2', 35.00, 135.00),
    ('01-PRE-030', 'DEMOLICION DE MOLDURAS PERIMETRALES EXISTENTES EN FACHADA POR MEDIOS MANUALES', 'LOTE', 1.00, 3500.00),
    ('02-ALB', 'ALBAÑILERIA Y RECUBRIMIENTOS', '', None, None),
    ('02-ALB-005', 'RESANE Y REPOSICION PUNTUAL DE PIEZAS DE ADOBE DANADAS CON MORTERO DE BARRO-CAL COMPATIBLE', 'M2', 6.00, 185.00),
    ('02-ALB-010', 'APLANADO REPELLADO Y AFINADO LISO EN MUROS DE ADOBE CON MORTERO CAL-ARENA PROP. 1:3', 'M2', 40.00, 340.00),
    ('02-ALB-020', 'APLICACION DE ESTUCO EN MUROS PERIMETRALES DE FACHADA SEGUN MUESTRA APROBADA', 'M2', 30.00, 180.00),
    ('02-ALB-030', 'CONSTRUCCION DE MURETE EN FACHADA A BASE DE BLOCK DE CONCRETO 15x20x40 CM, ALTURA HASTA 1.00 M', 'ML', 4.00, 795.00),
    ('02-ALB-035', 'SUMINISTRO, TENDIDO Y COMPACTACION DE MATERIAL DE BANCO PARA NIVELACION DE BASE EN COCHERA', 'M3', 2.50, 290.00),
    ('02-ALB-040', "FIRME DE CONCRETO F'C=200 KG/CM2 DE 10 CM EN COCHERA ARMADO CON MALLA 6x6-10/10", 'M2', 35.00, 445.00),
    ('03-PLA', 'PLAFONES', '', None, None),
    ('03-PLA-010', 'SUM. Y COLOC. DE PLAFON CORRIDO EN COCHERA A BASE DE PANEL DE YESO PARA EXTERIOR DE 12.7 MM', 'M2', 35.00, 520.00),
    ('03-PLA-020', 'DETALLADO DE CAJILLOS EXTERIORES CON PLAFON DE PANEL DE YESO, DESARROLLO HASTA 60 CM', 'ML', 15.00, 350.00),
    ('04-PIN', 'PINTURA', '', None, None),
    ('04-PIN-010', 'APLICACION DE PINTURA VINIL-ACRILICA PARA EXTERIOR A 2 MANOS SOBRE SELLADOR ACRILICO 5X1', 'M2', 90.00, 95.00),
    ('04-PIN-020', 'APLICACION DE SELLADOR Y PINTURA ELASTOMERICA TRANSPIRABLE A 2 MANOS EN LADRILLO APARENTE', 'M2', 15.00, 145.00),
    ('05-HOJ', 'HOJALATERIA', '', None, None),
    ('05-HOJ-010', 'REUBICACION DE CANALON DE LAMINA GALVANIZADA EXISTENTE SEGUN NUEVA POSICION', 'ML', 8.00, 260.00),
    ('06-LIM', 'LIMPIEZA Y ACARREOS', '', None, None),
    ('06-LIM-010', 'CARGA Y ACARREO EN CAMION DE MATERIAL PRODUCTO DE DEMOLICIONES A TIRO AUTORIZADO', 'M3', 8.00, 210.00),
    ('06-LIM-020', 'LIMPIEZA GRUESA Y FINA DE LAS AREAS INTERVENIDAS AL TERMINO DE LOS TRABAJOS', 'LOTE', 1.00, 2000.00),
]
ENCABEZADOS_OPUS = ['Clave', 'Concepto', 'Unidad', 'Cantidad', 'P.U.', 'Importe']

def filas_opus(con_precios=True):
    filas = []
    for cve, desc, uni, cant, pu in OPUS24:
        if cant is None:
            sub = sum(c * p for k, d, u, c, p in OPUS24 if c is not None and k.startswith(cve + '-'))
            filas.append([cve, desc, '', '', '', round(sub, 2) if con_precios else ''])
        elif con_precios:
            filas.append([cve, desc, uni, cant, pu, round(cant * pu, 2)])
        else:
            filas.append([cve, desc, uni, cant, '', ''])
    return filas

def escribir_opus(nombre, titulo, con_precios):
    filas = filas_opus(con_precios)
    wb = Workbook(); ws = wb.active; ws.title = 'Presupuesto'
    for _ in range(4): ws.append([])
    ws['B5'] = titulo; ws['B5'].font = Font(bold=True, size=12)     # OPUS: nombre del proyecto en B5
    ws.append([]); ws.append(ENCABEZADOS_OPUS)                       # encabezados en la fila 6
    for c in ws[6]: c.font = Font(bold=True)
    for f in filas: ws.append(f)
    if con_precios:
        ws.append([]); ws.append(['', 'TOTAL', '', '', '', round(sum(c * p for k, d, u, c, p in OPUS24 if c is not None), 2)])
    for col, w in zip('ABCDEF', [14, 70, 8, 12, 14, 16]): ws.column_dimensions[col].width = w
    wb.save(os.path.join(R, 'docs', 'plantillas', nombre + '.xlsx'))
    with io.open(os.path.join(R, 'scripts', 'qa', 'fixtures', nombre + '.csv'), 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f)
        for _ in range(4): w.writerow([])
        w.writerow(['', titulo]); w.writerow([]); w.writerow(ENCABEZADOS_OPUS)
        for fila in filas: w.writerow(fila)
    return len(filas)

n1 = escribir_opus('opus-24-ejemplo', 'REMODELACION DE FACHADA FRONTAL CASA AV. ORTIZ MENA', True)
n2 = escribir_opus('opus-24-concurso', 'ADAPTACION DE SUCURSAL CD. CUAUHTEMOC CHIHUAHUA CR 152', False)
print('plantillas OPUS 24 listas;', n1, 'renglones con precio y', n2, 'para cotizar')
